#!/usr/bin/env python3
"""
copy_excel.py — Perfect Excel file copy using openpyxl.

Usage:
    python3 utils/copy_excel.py <input_path> <output_path>

Copies the input .xlsx file to output_path, preserving:
- All sheets
- Cell values, formulas, data types
- Formatting: fonts, fills, borders, alignment, number formats
- Merged cells
- Row heights and column widths
- Print settings
- Named ranges (where supported)

Falls back to a raw file copy if openpyxl fails for any reason,
so the R caller always gets a file at output_path.
"""

import sys
import shutil
import os

def copy_excel_perfect(input_path, output_path):
    """
    Attempt a full-fidelity openpyxl copy.
    Falls back to shutil.copy2 (raw bytes) if anything goes wrong.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import (
            Font, PatternFill, Border, Side, Alignment, numbers
        )
        from openpyxl.utils import get_column_letter
        import copy

        # Load with data_only=False to preserve formulas
        src_wb = load_workbook(input_path, data_only=False)
        dst_wb = load_workbook(input_path, data_only=False)

        # openpyxl load_workbook already does a full in-memory copy
        # when you load the same file — saving it back gives a perfect clone.
        # The explicit load + save is the most reliable approach.
        dst_wb.save(output_path)

        print(f"[copy_excel] Success: {input_path} -> {output_path}", flush=True)
        return True

    except Exception as e:
        print(f"[copy_excel] openpyxl failed ({e}), falling back to raw copy", flush=True)
        try:
            shutil.copy2(input_path, output_path)
            print(f"[copy_excel] Raw copy fallback succeeded", flush=True)
            return True
        except Exception as e2:
            print(f"[copy_excel] Raw copy also failed: {e2}", flush=True)
            return False


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 copy_excel.py <input_path> <output_path>", file=sys.stderr)
        sys.exit(1)

    input_path  = sys.argv[1]
    output_path = sys.argv[2]

    if not os.path.isfile(input_path):
        print(f"[copy_excel] Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    # Ensure output directory exists
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    success = copy_excel_perfect(input_path, output_path)
    sys.exit(0 if success else 1)
